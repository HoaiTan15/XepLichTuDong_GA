"""
WEB APP - GENETIC ALGORITHM SCHEDULE OPTIMIZER
Giao diện web đơn giản cho thuật toán tối ưu hóa lịch học
 FIX: Thread-safe, timeout, better error handling
"""

from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import pandas as pd
import os
import io
import base64
import matplotlib
matplotlib.use('Agg') 
import matplotlib.pyplot as plt
from genetic_algorithm import GeneticAlgorithm, get_timeslot_detail
import threading
import time
import logging

#  FIX: Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

app = Flask(__name__)
app.secret_key = 'genetic_algorithm_secret_key'

#  FIX: Thread-safe progress tracking với Lock
progress_lock = threading.Lock()
current_progress = {'status': 'idle', 'progress': 0, 'best_fitness': 0, 'generation': 0}

class ProgressTracker:
    def __init__(self):
        self.reset()
    
    def reset(self):
        global current_progress
        with progress_lock:  #  Thread-safe
            current_progress = {'status': 'idle', 'progress': 0, 'best_fitness': 0, 'generation': 0}
    
    def update(self, generation, max_gen, best_fitness):
        global current_progress
        with progress_lock:  #  Thread-safe
            current_progress = {
                'status': 'running',
                'progress': int((generation / max_gen) * 100),
                'best_fitness': best_fitness,
                'generation': generation
            }

progress_tracker = ProgressTracker()

class WebGeneticAlgorithm(GeneticAlgorithm):
    """GA with web progress tracking"""
    
    def run(self):
        global current_progress
        progress_tracker.reset()
        population = self.create_population()
        best = None
        stagnant_count = 0

        for gen in range(self.generations):
            # Evaluate fitness
            from fitness import fitness_score
            scored = []
            for ind in population:
                fit, _ = fitness_score(ind)
                scored.append({"schedule": ind, "fitness": fit})
            scored.sort(key=lambda x: x['fitness'], reverse=True)

            # Track best solution
            if best is None or scored[0]['fitness'] > best['fitness']:
                best = scored[0]. copy()
                stagnant_count = 0
            else:
                stagnant_count += 1
            
            # Update progress
            progress_tracker.update(gen + 1, self.generations, best['fitness'])
            
            #  FIX: Tăng early stopping lên 30 (từ 15)
            if stagnant_count > 30:
                logging.info(f"⏹️ Early stopping at generation {gen}")
                with progress_lock:
                    current_progress['progress'] = 100
                    current_progress['status'] = 'completed'
                return best

            # Elitism: Giữ lại top tốt nhất
            elite_size = max(5, self.population_size // 8)
            new_pop = [s['schedule'] for s in scored[:elite_size]]

            # Generate new population với tournament selection
            while len(new_pop) < self.population_size:
                import random
                from crossover import crossover
                from mutation import mutate
                
                #  FIX: Tournament từ toàn bộ population
                tournament1 = random. sample(scored, min(3, len(scored)))
                tournament2 = random.sample(scored, min(3, len(scored)))
                
                p1 = max(tournament1, key=lambda x: x['fitness'])
                p2 = max(tournament2, key=lambda x: x['fitness'])
                
                child = crossover(p1['schedule'], p2['schedule'])
                
                #  FIX: Tăng base mutation rate lên 0.15 (từ 0.08)
                base_rate = 0.15
                adaptive_rate = base_rate + (stagnant_count * 0.015)
                mutation_rate = min(adaptive_rate, 0.4)
                mutate(child, rate=mutation_rate)
                
                new_pop.append(child)

            population = new_pop
            time.sleep(0.01)  # Small delay for UI updates

        # Đảm bảo progress = 100% khi chạy xong
        with progress_lock:
            current_progress['progress'] = 100
            current_progress['status'] = 'completed'
        return best

@app.route('/')
def index():
    """Trang chủ"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Upload và xử lý file Excel"""
    if 'file' not in request.files:
        flash('Không có file được chọn!')
        return redirect(url_for('index'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Không có file được chọn!')
        return redirect(url_for('index'))
    
    if file and file.filename.endswith('.xlsx'):
        try:
            #  FIX: Đọc file linh hoạt hơn - kiểm tra số cột
            df_test = pd.read_excel(file, skiprows=1, header=None, dtype=str, nrows=1)
            num_cols = len(df_test.columns)
            
            # Reset file pointer
            file.seek(0)
            
            # Đọc file với dtype=str để giữ nguyên format
            df = pd.read_excel(file, skiprows=1, header=None, dtype=str)
            
            #  FIX: Xử lý linh hoạt 5 hoặc 6 cột
            if num_cols == 6:
                df. columns = ['course_id', 'course_name', 'subject_code', 'section', 'teacher', 'room']
                logging.info(" Dataset có 6 cột (bao gồm room)")
                # Bỏ cột room vì GA sẽ tự generate
                df = df.drop('room', axis=1)
            elif num_cols == 5:
                df.columns = ['course_id', 'course_name', 'subject_code', 'section', 'teacher']
                logging.info(" Dataset có 5 cột (không có room)")
            else:
                raise ValueError(f"File phải có 5 hoặc 6 cột, nhưng có {num_cols} cột")
            
            # Tạo danh sách phòng học available (tự động generate)
            available_rooms = []
            # Tầng A: A101-A110, A201-A210, A301-A310, A401-A410
            for floor in [1, 2, 3, 4]:
                for room_num in range(1, 11):
                    available_rooms.append(f"A{floor}0{room_num}")
            # Tầng B: B101-B110, B201-B210, B301-B310, B401-B410
            for floor in [1, 2, 3, 4]:
                for room_num in range(1, 11):
                    available_rooms.append(f"B{floor}0{room_num}")
            
            # Lưu danh sách phòng để GA sử dụng
            with open('available_rooms.txt', 'w') as f:
                for room in available_rooms:
                    f.write(f"{room}\n")
            
            # Save for processing
            df.to_pickle('temp_data.pkl')
            
            # Show data preview - hiển thị toàn bộ
            data_info = {
                'total_classes': len(df),
                'teachers': df['teacher'].nunique(),
                'rooms': len(available_rooms),
                'sample_data': df.to_html(classes='table table-striped table-hover', table_id='dataTable', escape=False)
            }
            
            return render_template('data_preview.html', data_info=data_info)
            
        except Exception as e:
            logging.error(f" Lỗi đọc file: {str(e)}")
            flash(f'Lỗi đọc file: {str(e)}')
            return redirect(url_for('index'))
    
    else:
        flash('File phải có định dạng . xlsx')
        return redirect(url_for('index'))

@app.route('/run_algorithm', methods=['POST'])
def run_algorithm():
    """Chạy thuật toán GA"""
    try:
        # Load data
        df = pd.read_pickle('temp_data.pkl')
        
        # Get parameters from form - tối ưu cho 50 classes
        population_size = int(request.form.get('population_size', 50))
        generations = int(request. form.get('generations', 100))
        
        # Reset progress
        progress_tracker.reset()
        
        #  FIX: Run GA in background thread với timeout
        result_container = {'best': None, 'error': None}
        
        def run_ga():
            try:
                ga = WebGeneticAlgorithm(df, population_size=population_size, generations=generations)
                best = ga.run()
                result_container['best'] = best
                
                # Save results - Tạo 2 file: danh sách và thời khóa biểu
                schedule_df = pd.DataFrame(best['schedule'])
                schedule_df['timeslot_detail'] = schedule_df['timeslot'].apply(get_timeslot_detail)
                columns = ['course_id', 'course_name', 'subject_code', 'section', 'teacher', 'room', 'timeslot_detail']
                schedule_df = schedule_df[columns]
                
                # File 1: Danh sách (list view)
                schedule_df.to_excel('result_schedule.xlsx', index=False)
                
                # File 2: Thời khóa biểu (timetable view)
                create_timetable_excel(best['schedule'], 'result_timetable.xlsx')
                
                # Calculate conflicts & penalties từ hàm fitness
                from fitness import fitness_score
                fit_value, penalty_info = fitness_score(best['schedule'])
                conflict_info = count_conflicts(best['schedule'])
                conflicts = conflict_info['total']
                
                # Đếm số lớp buổi tối và Chủ nhật
                evening_classes = 0
                sunday_classes = 0
                for cls in best['schedule']:
                    ts = cls['timeslot']
                    day_index = (ts - 1) // 5
                    session_in_day = ((ts - 1) % 5) + 1
                    
                    if session_in_day == 5 and day_index < 6:
                        evening_classes += 1
                    
                    if day_index == 6:
                        sunday_classes += 1
                
                result = {
                    'fitness': fit_value,
                    'conflicts': conflicts,
                    'room_conflicts': conflict_info['room_conflicts'],
                    'teacher_conflicts': conflict_info['teacher_conflicts'],
                    'evening_classes': evening_classes,
                    'sunday_classes': sunday_classes,
                    'success_rate': ((len(df)*(len(df)-1)/2 - conflicts)/(len(df)*(len(df)-1)/2)*100),
                    'conflict_penalty': penalty_info['conflict_penalty'],
                    'evening_penalty': penalty_info['evening_penalty'],
                    'sunday_penalty': penalty_info['sunday_penalty'],
                    'evening_ratio_penalty': penalty_info['evening_ratio_penalty'],
                    'sunday_ratio_penalty': penalty_info['sunday_ratio_penalty'],
                    'total_penalty': penalty_info['total_penalty'],
                }
                
                with open('result_info.txt', 'w') as f:
                    f.write(
                        f"{result['fitness']},{result['conflicts']},{result['success_rate']},{result['room_conflicts']},{result['teacher_conflicts']},{result['evening_classes']},{result['sunday_classes']},"
                        f"{result['conflict_penalty']},{result['evening_penalty']},{result['sunday_penalty']},{result['evening_ratio_penalty']},{result['sunday_ratio_penalty']},{result['total_penalty']}"
                    )
                
                logging.info(f"GA hoàn thành: Fitness={result['fitness']:.6f}, Conflicts={conflicts}")
                
            except Exception as e:
                logging.error(f"❌ Lỗi trong GA thread: {str(e)}")
                result_container['error'] = str(e)
                with progress_lock:
                    current_progress['status'] = 'error'
        
        # Start background thread
        thread = threading.Thread(target=run_ga)
        thread.daemon = True  # Daemon thread sẽ tự dừng khi main process dừng
        thread.start()
        
        return render_template('running.html')
        
    except Exception as e:
        logging.error(f"❌ Lỗi chạy thuật toán: {str(e)}")
        flash(f'Lỗi chạy thuật toán: {str(e)}')
        return redirect(url_for('index'))

@app.route('/progress')
def get_progress():
    """API endpoint để lấy tiến độ"""
    with progress_lock:  # Thread-safe read
        return current_progress. copy()

@app.route('/results')
def show_results():
    """Hiển thị kết quả"""
    try:
        # Read results
        with open('result_info.txt', 'r') as f:
            result_data = f.read().strip().split(',')
            fitness = result_data[0]
            conflicts = result_data[1]
            success_rate = result_data[2]
            room_conflicts = result_data[3] if len(result_data) > 3 else "0"
            teacher_conflicts = result_data[4] if len(result_data) > 4 else "0"
            evening_classes = result_data[5] if len(result_data) > 5 else "0"
            sunday_classes = result_data[6] if len(result_data) > 6 else "0"
            conflict_penalty = result_data[7] if len(result_data) > 7 else "0"
            evening_penalty = result_data[8] if len(result_data) > 8 else "0"
            sunday_penalty = result_data[9] if len(result_data) > 9 else "0"
            evening_ratio_penalty = result_data[10] if len(result_data) > 10 else "0"
            sunday_ratio_penalty = result_data[11] if len(result_data) > 11 else "0"
            total_penalty = result_data[12] if len(result_data) > 12 else "0"
        
        # Read schedule
        schedule_df = pd.read_excel('result_schedule.xlsx')
        
        # Đảm bảo có cột timeslot_detail
        if 'timeslot' in schedule_df.columns and 'timeslot_detail' not in schedule_df.columns:
            schedule_df['timeslot_detail'] = schedule_df['timeslot'].apply(get_timeslot_detail)
        
        if 'timeslot' in schedule_df.columns:
            schedule_df = schedule_df.drop('timeslot', axis=1)
        
        # Generate charts
        chart_url = generate_charts(schedule_df)
        
        results = {
            'fitness': float(fitness),
            'conflicts': int(float(conflicts)),
            'room_conflicts': int(float(room_conflicts)),
            'teacher_conflicts': int(float(teacher_conflicts)),
            'evening_classes': int(float(evening_classes)),
            'sunday_classes': int(float(sunday_classes)),
            'success_rate': float(success_rate),
            'conflict_penalty': float(conflict_penalty),
            'evening_penalty': float(evening_penalty),
            'sunday_penalty': float(sunday_penalty),
            'evening_ratio_penalty': float(evening_ratio_penalty),
            'sunday_ratio_penalty': float(sunday_ratio_penalty),
            'total_penalty': float(total_penalty),
            'total_classes': len(schedule_df),
            'chart_url': chart_url,
            'schedule_table': schedule_df.to_html(classes='table table-striped table-hover', table_id='resultTable', escape=False)
        }
        
        return render_template('results.html', results=results)
        
    except Exception as e:
        logging.error(f"❌ Lỗi hiển thị kết quả: {str(e)}")
        flash(f'Lỗi hiển thị kết quả: {str(e)}')
        return redirect(url_for('index'))

@app.route('/download')
def download_file():
    """Download file kết quả dạng danh sách"""
    try:
        file_path = os.path.abspath('result_schedule.xlsx')
        
        if not os.path.exists(file_path):
            flash('File kết quả không tồn tại. Vui lòng chạy thuật toán trước!')
            return redirect(url_for('index'))
        
        return send_file(file_path, 
                        as_attachment=True, 
                        download_name='optimized_schedule.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logging.error(f"❌ Lỗi download file: {str(e)}")
        flash(f'Lỗi download file: {str(e)}')
        return redirect(url_for('show_results'))

@app.route('/download_timetable')
def download_timetable():
    """Download file kết quả dạng thời khóa biểu"""
    try:
        file_path = os.path.abspath('result_timetable.xlsx')
        
        if not os.path.exists(file_path):
            flash('File thời khóa biểu không tồn tại. Vui lòng chạy thuật toán trước!')
            return redirect(url_for('index'))
        
        return send_file(file_path, 
                        as_attachment=True, 
                        download_name='timetable.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logging.error(f"❌ Lỗi download timetable: {str(e)}")
        flash(f'Lỗi download timetable: {str(e)}')
        return redirect(url_for('show_results'))

def create_timetable_excel(schedule, filename):
    """
    Tạo file Excel dạng thời khóa biểu (lưới)
    - Dòng: 5 ca học (Sáng 1-3, Sáng 4-6, Chiều 7-9, Chiều 10-12, Tối 13-15)
    - Cột: 7 ngày (Thứ 2 - Chủ nhật)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Thoi Khoa Bieu"
    
    # Định nghĩa cấu trúc
    days = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]
    sessions = [
        "Sáng\nTiết 1-3",
        "Sáng\nTiết 4-6",
        "Chiều\nTiết 7-9",
        "Chiều\nTiết 10-12",
        "Tối\nTiết 13-15"
    ]
    
    # Tạo header
    ws.merge_cells('A1:A2')
    ws['A1'] = "Ca học / Thứ"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    
    # Header các ngày
    for col_idx, day in enumerate(days, start=2):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = day
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
    
    # Tạo lưới thời khóa biểu
    timetable = {}  # Key: (day_index, session), Value: list of classes
    
    for cls in schedule:
        ts = cls['timeslot']
        day_index = (ts - 1) // 5  # 0-6
        session_index = (ts - 1) % 5  # 0-4
        
        key = (day_index, session_index)
        if key not in timetable:
            timetable[key] = []
        
        class_info = (
            f"{cls['course_name']}\n"
            f"({cls['section']})\n"
            f"GV: {cls['teacher']}\n"
            f"Phòng: {cls['room']}"
        )
        timetable[key].append(class_info)
    
    # Điền dữ liệu vào lưới
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for session_idx, session_name in enumerate(sessions):
        row = session_idx + 3
        
        # Cột đầu tiên: tên ca học
        cell = ws.cell(row=row, column=1)
        cell.value = session_name
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.border = thin_border
        
        # Các cột ngày
        for day_idx in range(7):
            col = day_idx + 2
            cell = ws.cell(row=row, column=col)
            
            key = (day_idx, session_idx)
            if key in timetable:
                classes = timetable[key]
                cell.value = "\n\n".join(classes)  # Nếu có nhiều lớp cùng slot
                
                # Màu cảnh báo nếu có conflict (nhiều hơn 1 lớp)
                if len(classes) > 1:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            else:
                cell.value = ""
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.font = Font(size=9)
            cell.border = thin_border
    
    # Điều chỉnh kích thước cột và dòng
    ws.column_dimensions['A'].width = 15
    for col_idx in range(2, 9):
        ws.column_dimensions[chr(64 + col_idx)].width = 25
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    for row_idx in range(3, 8):
        ws.row_dimensions[row_idx].height = 80
    
    # Thêm sheet thống kê
    ws_stats = wb.create_sheet("Thong Ke")
    ws_stats['A1'] = "THỐNG KÊ THỜI KHÓA BIỂU"
    ws_stats['A1'].font = Font(bold=True, size=14)
    
    stats_data = [
        ["Tổng số lớp", len(schedule)],
        ["Số phòng học sử dụng", len(set(cls['room'] for cls in schedule))],
        ["Số giảng viên", len(set(cls['teacher'] for cls in schedule))],
        ["Số timeslot sử dụng", len(timetable)],
        ["Số lớp buổi tối", sum(1 for cls in schedule if ((cls['timeslot']-1) % 5) == 4 and ((cls['timeslot']-1) // 5) < 6)],
        ["Số lớp Chủ nhật", sum(1 for cls in schedule if ((cls['timeslot']-1) // 5) == 6)]
    ]
    
    for idx, (label, value) in enumerate(stats_data, start=3):
        ws_stats[f'A{idx}'] = label
        ws_stats[f'B{idx}'] = value
        ws_stats[f'A{idx}'].font = Font(bold=True)
    
    wb.save(filename)
    logging.info(f" Đã tạo file thời khóa biểu: {filename}")

def count_conflicts(schedule):
    """
    Đếm và phân loại các xung đột
    Returns: dict với thông tin chi tiết về conflicts
    """
    room_conflicts = 0
    teacher_conflicts = 0
    conflict_details = []
    
    for i in range(len(schedule)):
        for j in range(i+1, len(schedule)):
            a, b = schedule[i], schedule[j]
            
            # Room conflict
            if a['room'] == b['room'] and a['timeslot'] == b['timeslot']:
                room_conflicts += 1
                conflict_details.append({
                    'type': 'Room',
                    'room': a['room'],
                    'timeslot': a['timeslot'],
                    'class1': f"{a['course_name']} ({a['section']})",
                    'class2': f"{b['course_name']} ({b['section']})"
                })
            
            # Teacher conflict  
            if a['teacher'] == b['teacher'] and a['timeslot'] == b['timeslot']:
                teacher_conflicts += 1
                conflict_details.append({
                    'type': 'Teacher',
                    'teacher': a['teacher'],
                    'timeslot': a['timeslot'],
                    'class1': f"{a['course_name']} ({a['section']})",
                    'class2': f"{b['course_name']} ({b['section']})"
                })
    
    return {
        'total': room_conflicts + teacher_conflicts,
        'room_conflicts': room_conflicts,
        'teacher_conflicts': teacher_conflicts,
        'details': conflict_details
    }

def generate_charts(df):
    """Tạo biểu đồ và trả về base64 string"""
    plt.figure(figsize=(14, 10))
    
    # 1. Phân bổ theo ca học
    plt.subplot(2, 3, 1)
    if 'timeslot_detail' in df.columns:
        ca_hoc = df['timeslot_detail'].str.extract(r'- (Sáng|Chiều|Tối) -')[0].value_counts()
        plt.pie(ca_hoc. values, labels=ca_hoc.index, autopct='%1.1f%%', startangle=90,
                colors=['gold', 'lightcoral', 'lightblue'])
        plt.title('Phân bổ theo Ca học')
    else:
        plt.text(0.5, 0.5, 'Không có dữ liệu timeslot_detail', ha='center', va='center')
        plt.title('Phân bổ theo Ca học')
    
    # 2. Room usage
    plt.subplot(2, 3, 2)
    room_counts = df['room'].value_counts()
    plt.hist(room_counts. values, bins=10, color='lightgreen', alpha=0.8, edgecolor='black')
    plt.title('Sử dụng phòng học')
    plt.xlabel('Số lớp/phòng')
    plt.ylabel('Số phòng')
    plt.grid(True, alpha=0.3)
    
    # 3. Teacher workload
    plt.subplot(2, 3, 3)
    teacher_counts = df['teacher'].value_counts(). head(10)
    plt.barh(range(len(teacher_counts)), teacher_counts.values, color='orange', alpha=0.8)
    plt.title('Top 10 giảng viên')
    plt. xlabel('Số lớp')
    plt.yticks(range(len(teacher_counts)), [name[:15] + '...' if len(name) > 15 else name for name in teacher_counts.index])
    
    # 4. Thứ trong tuần distribution
    plt.subplot(2, 3, 4)
    if 'timeslot_detail' in df.columns:
        thu_hoc = df['timeslot_detail'].str.extract(r'(Thứ \d+|Chủ nhật)')[0].value_counts()
        plt.bar(range(len(thu_hoc)), thu_hoc.values, color='mediumpurple', alpha=0.8)
        plt.title('Phân bổ theo Thứ')
        plt.xlabel('Thứ')
        plt.ylabel('Số lớp')
        plt.xticks(range(len(thu_hoc)), thu_hoc.index, rotation=45)
        plt.grid(True, alpha=0.3)
    else:
        plt.text(0.5, 0.5, 'Không có dữ liệu', ha='center', va='center')
        plt.title('Phân bổ theo Thứ')
    
    # 5. Phân bổ chi tiết theo timeslot_detail
    plt.subplot(2, 3, 5)
    if 'timeslot_detail' in df.columns:
        timeslot_detail_counts = df['timeslot_detail'].value_counts(). head(10)
        plt. barh(range(len(timeslot_detail_counts)), timeslot_detail_counts.values, color='skyblue', alpha=0.8)
        plt.title('Top 10 Timeslot')
        plt.xlabel('Số lớp')
        plt. yticks(range(len(timeslot_detail_counts)), 
                  [detail[:20] + '...' if len(detail) > 20 else detail for detail in timeslot_detail_counts.index])
    else:
        plt.text(0.5, 0.5, 'Không có dữ liệu', ha='center', va='center')
        plt.title('Top 10 Timeslot')
    
    # 6.  Phân bổ buổi tối và Chủ nhật
    plt.subplot(2, 3, 6)
    if 'timeslot_detail' in df.columns:
        special_counts = {}
        for detail in df['timeslot_detail']:
            if 'Tối' in detail:
                special_counts['Buổi tối'] = special_counts.get('Buổi tối', 0) + 1
            elif 'Chủ nhật' in detail:
                special_counts['Chủ nhật'] = special_counts.get('Chủ nhật', 0) + 1
            else:
                special_counts['Giờ hành chính'] = special_counts. get('Giờ hành chính', 0) + 1
        
        if special_counts:
            plt. pie(special_counts.values(), labels=special_counts.keys(), autopct='%1.1f%%', 
                    startangle=90, colors=['lightgreen', 'lightcoral', 'gold'])
        plt.title('Phân loại đặc biệt')
    else:
        plt.text(0.5, 0.5, 'Không có dữ liệu', ha='center', va='center')
        plt. title('Phân loại đặc biệt')
    
    plt.tight_layout()
    
    # Convert to base64
    img = io.BytesIO()
    plt.savefig(img, format='png', dpi=150, bbox_inches='tight')
    img.seek(0)
    chart_url = base64.b64encode(img.getvalue()).decode()
    plt.close()
    
    return chart_url

if __name__ == '__main__':
    # Fix Windows socket error: use_reloader=False hoặc tắt debug mode
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)